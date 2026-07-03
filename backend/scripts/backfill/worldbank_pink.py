"""World Bank Pink Sheet → public.historical_prices.

Monthly commodity prices back to 1960. Public XLSX, no auth required.
Requires `openpyxl` (already in the backend requirements).

Usage:
    python -m backend.scripts.backfill.worldbank_pink
"""

from __future__ import annotations

import argparse
import datetime as dt
import io
import logging
import tempfile
import urllib.request
from typing import Any

from .common import batch_upsert, get_supabase, setup_logging, write_cursor

logger = logging.getLogger("backfill.worldbank_pink")

SOURCE = "worldbank_pink"
_URL = (
    "https://thedocs.worldbank.org/en/doc/18675909112009424-0050022022/"
    "original/CMO-Historical-Data-Monthly.xlsx"
)

# Pink Sheet sheet name is "Monthly Prices". Column names use spaces + newlines
# in the actual file — we look for these substrings when scanning headers.
_COLUMN_MAP: dict[str, tuple[str, str]] = {
    # workbook column key substring → (canonical commodity, unit)
    "crude, brent":       ("brent", "usd_per_barrel"),
    "crude, wti":         ("wti", "usd_per_barrel"),
    "crude, dubai":       ("dubai", "usd_per_barrel"),
    "natural gas, us":    ("gas", "usd_per_mmbtu"),
    "gold":               ("gold", "usd_per_troy_oz"),
    "silver":             ("silver", "usd_per_troy_oz"),
    "copper":             ("copper", "usd_per_metric_ton"),
    "wheat, us hrw":      ("wheat", "usd_per_metric_ton"),
    "maize":              ("corn", "usd_per_metric_ton"),
    "soybeans":           ("soybean", "usd_per_metric_ton"),
    "coffee, arabica":    ("coffee", "usd_per_kg"),
    "cocoa":              ("cocoa", "usd_per_kg"),
    "sugar, world":       ("sugar", "usd_per_kg"),
    "cotton, a index":    ("cotton", "usd_per_kg"),
}


def _download() -> bytes:
    logger.debug("fetching %s", _URL)
    with urllib.request.urlopen(_URL, timeout=180) as resp:
        return resp.read()


def _parse_xlsx(xlsx_bytes: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        raise RuntimeError("openpyxl required — pip install openpyxl")

    rows_out: list[dict[str, Any]] = []
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
        tmp.write(xlsx_bytes)
        tmp.flush()
        wb = load_workbook(tmp.name, data_only=True, read_only=True)
        try:
            ws = wb["Monthly Prices"]
        except KeyError:
            # Newer files rename the sheet occasionally.
            ws = wb[wb.sheetnames[0]]

        # Header rows on Pink Sheet span two rows: line up the commodity name
        # in row 5-ish. We resolve by searching for column headers containing
        # our target substrings across the first 10 rows.
        header_map: dict[int, tuple[str, str]] = {}
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for col_idx, cell in enumerate(row):
                if not isinstance(cell, str):
                    continue
                cell_lc = cell.lower().strip()
                for token, (commodity, unit) in _COLUMN_MAP.items():
                    if token in cell_lc and col_idx not in header_map:
                        header_map[col_idx] = (commodity, unit)

        if not header_map:
            logger.warning("worldbank_pink: no matching columns found in header rows")
            return []

        # Data starts around row 7. First column is the date label
        # ("1960M01" style). Walk from row 7 onward.
        for row in ws.iter_rows(min_row=7, values_only=True):
            date_label = row[0]
            observed_on = _parse_date_label(date_label)
            if observed_on is None:
                continue
            for col_idx, (commodity, unit) in header_map.items():
                if col_idx >= len(row):
                    continue
                price = row[col_idx]
                if price is None:
                    continue
                try:
                    price_f = float(price)
                except (TypeError, ValueError):
                    continue
                rows_out.append(
                    {
                        "source": SOURCE,
                        "commodity": commodity,
                        "observed_on": observed_on.isoformat(),
                        "price": price_f,
                        "unit": unit,
                        "frequency": "monthly",
                    }
                )
    return rows_out


def _parse_date_label(label: Any) -> dt.date | None:
    if not isinstance(label, str):
        return None
    label = label.strip()
    # "1960M01" → 1960-01-01
    if "M" in label:
        try:
            year_s, month_s = label.split("M")
            return dt.date(int(year_s), int(month_s), 1)
        except (ValueError, IndexError):
            return None
    return None


def backfill(supabase, *, since: dt.date | None = None, until: dt.date | None = None) -> int:
    try:
        xlsx_bytes = _download()
    except Exception as exc:  # noqa: BLE001
        logger.warning("worldbank_pink: download failed (%s)", exc)
        return 0

    rows = _parse_xlsx(xlsx_bytes)
    if since or until:
        rows = [
            r for r in rows
            if (since is None or r["observed_on"] >= since.isoformat())
            and (until is None or r["observed_on"] < until.isoformat())
        ]

    upserted = batch_upsert(
        supabase, "historical_prices", rows, on_conflict="source,commodity,observed_on"
    )
    logger.info("worldbank_pink: %d rows (upserted %d)", len(rows), upserted)
    if rows:
        write_cursor(supabase, SOURCE, "observed_on", max(r["observed_on"] for r in rows), upserted)
    return upserted


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--since", help="Optional ISO date filter (inclusive)")
    ap.add_argument("--until", help="Optional ISO date filter (exclusive)")
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    setup_logging(args.verbose)
    supabase = get_supabase()
    since = dt.date.fromisoformat(args.since) if args.since else None
    until = dt.date.fromisoformat(args.until) if args.until else None
    total = backfill(supabase, since=since, until=until)
    print(f"worldbank_pink: {total} rows ingested")


if __name__ == "__main__":
    main()
